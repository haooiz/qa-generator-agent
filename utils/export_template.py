"""엑셀 출력용 템플릿 생성·관리.

서식(열 너비, 병합, 글꼴 등)은 templates/qa_export_template.xlsx 에서
Excel로 직접 수정할 수 있습니다. 앱은 이 템플릿에 데이터만 채웁니다.
"""

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TEMPLATE_DIR = Path(__file__).resolve().parent.parent / "templates"
TEMPLATE_PATH = TEMPLATE_DIR / "qa_export_template.xlsx"

HEADER_ROW = 4
DATA_START_ROW = 5
CONTENT_COL_END = 9
COL_A_WIDTH = 130  # Excel 열 너비 (문자 단위, Excel에서 보이는 값)

META_LABELS = ("문서명", "핵심 요약", "추천 복습 영역")
TABLE_HEADERS = (
    "번호",
    "유형",
    "문제",
    "선택지",
    "정답",
    "해설",
    "내 답변",
    "채점결과",
    "점수",
)
OTHER_COL_WIDTHS = (12, 40, 30, 25, 30, 25, 10, 8)


def _thin_border() -> Border:
    side = Side(style="thin", color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def create_export_template(path: Path | None = None) -> Path:
    """서식이 적용된 엑셀 템플릿 파일을 생성합니다."""
    path = path or TEMPLATE_PATH
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "질의응답"

    label_font = Font(bold=True, size=11)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    wrap_top = Alignment(wrap_text=True, vertical="top", horizontal="left")
    label_align = Alignment(vertical="center", horizontal="center", wrap_text=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = _thin_border()

    ws.column_dimensions["A"].width = COL_A_WIDTH
    for col_idx, width in enumerate(OTHER_COL_WIDTHS, 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row in (1, 2, 3):
        ws.row_dimensions[row].height = 48
        label_cell = ws.cell(row=row, column=1, value=META_LABELS[row - 1])
        label_cell.font = label_font
        label_cell.alignment = label_align
        label_cell.border = border

        content_cell = ws.cell(row=row, column=2, value="")
        content_cell.alignment = wrap_top
        content_cell.border = border
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=CONTENT_COL_END)

    for col, header in enumerate(TABLE_HEADERS, 1):
        cell = ws.cell(row=HEADER_ROW, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    ws.row_dimensions[HEADER_ROW].height = 22
    ws.freeze_panes = f"A{DATA_START_ROW}"

    wb.save(path)
    return path


def load_export_template() -> Workbook:
    """템플릿을 불러오거나 없으면 새로 만듭니다."""
    if not TEMPLATE_PATH.exists():
        create_export_template()
    return load_workbook(TEMPLATE_PATH)


def apply_data_row_style(ws, row: int) -> None:
    """데이터 행에 테두리·줄바꿈 스타일을 적용합니다."""
    border = _thin_border()
    align = Alignment(wrap_text=True, vertical="top", horizontal="left")
    center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
    for col in range(1, CONTENT_COL_END + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = border
        cell.alignment = center_align if col in (1, 2, 8, 9) else align
